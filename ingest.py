from openpyxl import load_workbook
from datetime import date, datetime
from models import Trade


REQUIRED_COLUMNS = {"Entry Date", "Trade P&L ($)", "Trade Return (%)", "R-Multiple"}


def read_trades(file_path: str) -> list[Trade]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows()
    header_row = next(rows)
    headers = {cell.value.strip(): idx for idx, cell in enumerate(header_row) if cell.value}

    missing = REQUIRED_COLUMNS - headers.keys()
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    col_date = headers["Entry Date"]
    col_pnl = headers["Trade P&L ($)"]
    col_ret = headers["Trade Return (%)"]
    col_r = headers["R-Multiple"]

    trades = []
    skipped = 0
    total = 0

    for row in rows:
        total += 1
        try:
            raw_date = row[col_date].value
            pnl = row[col_pnl].value
            ret = row[col_ret].value
            r = row[col_r].value

            if any(v is None for v in (raw_date, pnl, ret, r)):
                skipped += 1
                continue

            if isinstance(raw_date, datetime):
                entry_date = raw_date.date()
            elif isinstance(raw_date, date):
                entry_date = raw_date
            else:
                entry_date = date.fromisoformat(str(raw_date))

            trades.append(Trade(
                entry_date=entry_date,
                trade_pnl=float(pnl),
                trade_return_percent=float(ret),
                r_multiple=float(r),
            ))
        except (ValueError, TypeError, IndexError):
            skipped += 1

    wb.close()

    if skipped > 0:
        print(f"[warn] Skipped {skipped} of {total} rows due to missing or malformed data")

    trades.sort(key=lambda t: t.entry_date)
    return trades
